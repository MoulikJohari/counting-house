import csv
import io
import math
import re
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import AuditLog, Expense, Invoice, InvoicePayment, PurchaseOrder, RecurringExpense, User


def num(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


_DATE_FORMATS = (
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d.%m.%Y",
    "%d/%m/%y",
    "%d-%b-%Y",
    "%d-%b-%y",
    "%d %b %Y",
    "%m/%d/%Y",
)


def parse_date(value: str | date | None) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"invalid date '{text}'")


def parse_amount(value: Any, field: str = "amount") -> float:
    """Strict numeric parser for imports: blank -> 0.0, non-numeric junk -> ValueError."""
    if value is None:
        return 0.0
    if isinstance(value, bool):
        raise ValueError(f"invalid {field} '{value}'")
    if isinstance(value, (int, float)):
        if not math.isfinite(value):
            raise ValueError(f"invalid {field} '{value}'")
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    cleaned = re.sub(r"^(rs\.?|inr)\s*", "", text, flags=re.IGNORECASE)
    cleaned = cleaned.replace(",", "").replace("₹", "").replace("%", "").strip()
    try:
        result = float(cleaned)
    except ValueError:
        raise ValueError(f"invalid {field} '{text}'") from None
    if not math.isfinite(result):
        raise ValueError(f"invalid {field} '{text}'")
    return result


def _is_blank_row(values: Any) -> bool:
    if values is None:
        return True
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_csv_rows(content: str) -> list[tuple[int, dict[str, str]]]:
    """Return (line_number, row) pairs. Ghost/blank rows are dropped silently."""
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for raw in reader:
        values = {_normalize_header(k): (v or "").strip() for k, v in raw.items() if k is not None}
        if all(v == "" for v in values.values()):
            continue
        rows.append((reader.line_num, values))
    return rows


def parse_xlsx_rows(content: bytes) -> list[tuple[int, dict[str, str]]]:
    """Return (worksheet_row_number, row) pairs from the first sheet. Ghost/blank rows are dropped silently."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        if ws is None:
            return []
        rows_iter = ws.iter_rows(values_only=True)
        keys: list[str] = []
        header_row_num = 0
        for row_num, raw in enumerate(rows_iter, start=1):
            if _is_blank_row(raw):
                continue
            keys = [_normalize_header(h) for h in raw]
            header_row_num = row_num
            break
        if not keys:
            return []
        rows = []
        for row_num, raw in enumerate(rows_iter, start=header_row_num + 1):
            if _is_blank_row(raw):
                continue
            rows.append((row_num, {key: _cell_text(v) for key, v in zip(keys, raw)}))
        return rows
    finally:
        wb.close()


def get_field(row: dict[str, str], *names: str) -> str | None:
    for name in names:
        key = name.strip().lower().replace(" ", "_")
        value = row.get(key)
        if value:
            return value
    return None


def inv_calc(invoice: Invoice | dict, today: date | None = None) -> dict:
    today = today or date.today()
    if isinstance(invoice, dict):
        taxable = num(invoice.get("amount"))
        gst_rate = num(invoice.get("gst_rate"))
        tds_rate = num(invoice.get("tds_rate"))
        payments = invoice.get("payments") or []
        due_date = invoice.get("due_date")
    else:
        taxable = num(invoice.amount)
        gst_rate = num(invoice.gst_rate)
        tds_rate = num(invoice.tds_rate)
        payments = invoice.payments or []
        due_date = invoice.due_date

    gst = taxable * gst_rate / 100
    gross = taxable + gst
    tds = taxable * tds_rate / 100
    net = gross - tds
    collected = sum(num(p.amount if hasattr(p, "amount") else p.get("amount")) for p in payments)
    balance = max(0, net - collected)
    status = "unpaid"
    if net > 0 and collected >= net - 0.5:
        status = "paid"
    elif collected > 0.5:
        status = "partial"
    overdue = status != "paid" and due_date and due_date < today
    return {
        "taxable": taxable,
        "gst": gst,
        "gross": gross,
        "tds": tds,
        "net": net,
        "collected": collected,
        "balance": balance,
        "status": status,
        "overdue": overdue,
    }


def po_calc(po: PurchaseOrder, invoices: list[Invoice]) -> dict:
    val = num(po.amount)
    inv = sum(num(i.amount) for i in invoices if i.po_id == po.id)
    remaining = max(0, val - inv)
    pct = min(100, round(inv / val * 100)) if val else 0
    return {"val": val, "inv": inv, "remaining": remaining, "pct": pct}


def generate_recurring(db: Session) -> None:
    now = datetime.utcnow()
    cur_key = f"{now.year}-{now.month:02d}"
    recurring_items = db.query(RecurringExpense).filter(RecurringExpense.active.is_(True)).all()
    for r in recurring_items:
        if not r.start:
            continue
        y, m = map(int, r.start.split("-"))
        while True:
            key = f"{y}-{m:02d}"
            if key > cur_key:
                break
            exists = (
                db.query(Expense)
                .filter(Expense.recurring_id == r.id, Expense.rkey == key)
                .first()
            )
            if not exists:
                day = min(max(r.day or 1, 1), 28)
                db.add(
                    Expense(
                        category=r.category,
                        amount=r.amount,
                        vendor=r.label,
                        notes="Recurring",
                        date=date(y, m, day),
                        recurring_id=r.id,
                        rkey=key,
                    )
                )
            m += 1
            if m > 12:
                m = 1
                y += 1
    db.commit()


class AuditService:
    @staticmethod
    def log(
        db: Session,
        *,
        user: User | None,
        action: str,
        entity_type: str,
        entity_id: str | None,
        summary: str,
        changes: dict | None = None,
    ) -> AuditLog:
        entry = AuditLog(
            user_id=user.id if user else None,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            summary=summary,
            changes=changes,
        )
        db.add(entry)
        db.flush()
        return entry


def invoice_to_dict(invoice: Invoice) -> dict:
    return {
        "id": str(invoice.id),
        "company": invoice.company,
        "ref": invoice.ref,
        "date": invoice.date.isoformat(),
        "due_date": invoice.due_date.isoformat() if invoice.due_date else None,
        "po_id": str(invoice.po_id) if invoice.po_id else None,
        "amount": invoice.amount,
        "gst_rate": invoice.gst_rate,
        "tds_rate": invoice.tds_rate,
        "notes": invoice.notes,
        "payments": [
            {"id": str(p.id), "date": p.date.isoformat(), "amount": p.amount}
            for p in invoice.payments
        ],
    }


def po_to_dict(po: PurchaseOrder) -> dict:
    return {
        "id": str(po.id),
        "company": po.company,
        "ref": po.ref,
        "date": po.date.isoformat(),
        "amount": po.amount,
        "gst_rate": po.gst_rate,
        "notes": po.notes,
    }


def expense_to_dict(expense: Expense) -> dict:
    return {
        "id": str(expense.id),
        "category": expense.category,
        "date": expense.date.isoformat(),
        "amount": expense.amount,
        "vendor": expense.vendor,
        "notes": expense.notes,
        "recurring_id": str(expense.recurring_id) if expense.recurring_id else None,
        "rkey": expense.rkey,
    }
